from pathlib import Path

from openpyxl import load_workbook


def get_sheet_names(file_path: Path) -> list[str]:
    workbook = load_workbook(file_path, read_only=True)
    return workbook.sheetnames
