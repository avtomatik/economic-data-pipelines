from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.paths import DATA_DIR


def get_excel_sheet_names(file_path: Path) -> list[str]:
    workbook = load_workbook(file_path, read_only=True)
    return workbook.sheetnames


def read_excel_sheet(file_path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(file_path, sheet_name=sheet_name)


def main() -> None:
    dataset_files = list(DATA_DIR.glob("dataset_che_*.xls"))

    if not dataset_files:
        return

    for file_path in dataset_files:
        sheet_names = get_excel_sheet_names(file_path)

        for sheet_name in sheet_names:
            df = read_excel_sheet(file_path, sheet_name)
            print(df.head())
